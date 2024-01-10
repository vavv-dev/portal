from collections import OrderedDict

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from apps.ncs.models import NcsClassification


class Command(BaseCommand):
    help = "Load ncs classification from an Excel file"

    def add_arguments(self, parser):
        parser.add_argument("excel_file", type=str, help="Path to the Excel file")

    def handle(self, *args, **options):
        excel_file = options["excel_file"]
        wb = load_workbook(excel_file)

        # do not trigger signals
        NcsClassification.objects.all().delete()

        # distinct new NcsClassification
        distinct = OrderedDict()

        for row in wb.worksheets[0].iter_rows(min_row=2, values_only=True):
            pk = f"{row[0]}{row[2]}{row[4]}"

            class1_code, class1_name, class2_code, class2_name, class3_code, class3_name = row[0:6]

            classification = NcsClassification(
                pk=pk,
                class1_code=class1_code,
                class1_name=class1_name,
                class2_code=class2_code,
                class2_name=class2_name,
                class3_code=class3_code,
                class3_name=class3_name,
            )
            distinct[pk] = classification

        NcsClassification.objects.bulk_create(distinct.values())

        self.stdout.write(
            self.style.SUCCESS(
                f"Data loaded successfully from {excel_file} to NcsClassification model"
            )
        )
