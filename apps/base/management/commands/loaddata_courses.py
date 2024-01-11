from django.core.management.base import BaseCommand
from django.utils.text import slugify
from openpyxl import load_workbook

from apps.base.models import Category
from apps.course.models import Course, CourseHome
from apps.flex.models import Flex, FlexHome
from apps.program.models import Program, ProgramHome


class Command(BaseCommand):
    help = "Load category and course information from an Excel file"

    def add_arguments(self, parser):
        parser.add_argument("excel_file", type=str, help="Path to the Excel file")

    def create_or_get_category(self, name, level, parent=None):
        category = Category.objects.filter(name=name, tn_level=level, tn_parent=parent).last()
        if not category:
            category = Category.objects.create(name=name, tn_level=level, tn_parent=parent)
        return category

    def handle(self, *args, **options):
        excel_file = options["excel_file"]
        wb = load_workbook(excel_file)

        # course home
        course_home = CourseHome.objects.last()
        if not course_home:
            raise ValueError("Course Home not exists")

        # program home
        program_home = ProgramHome.objects.last()
        if not program_home:
            raise ValueError("Program Home not exists")

        # flex home
        flex_home = FlexHome.objects.last()
        if not flex_home:
            raise ValueError("Flex Home not exists")

        def handle_duplicate_slug(model, slug):
            if model.objects.filter(slug=slug):
                return f"{slug}-1"
            return slug

        for row in wb.worksheets[0].iter_rows(min_row=2, values_only=True):
            title, level1, level2, level3 = row[1:5]

            if not title:
                continue

            slug = slugify(str(title), allow_unicode=True)

            # create category
            category = self.create_or_get_category(level1, 1)
            if level2:
                category = self.create_or_get_category(level2, 2, category)
                if level3:
                    category = self.create_or_get_category(level3, 3, category)

            # check course
            course = None
            course = Course.objects.filter(title=title, categories__name=category.name).last()

            if not course:
                # fix duplicate course
                slug = handle_duplicate_slug(Course, slug)

                course = Course(title=title, slug=slug, allow_rating=True, allow_comments=True)
                course_home.add_child(instance=course)

                # add category
                course.categories.add(category)

                course.save()
                course.save_revision().publish()

            # program
            if not Program.objects.filter(title=title, categories__name=category.name).exists():
                # fix duplicate program
                slug = handle_duplicate_slug(Program, slug)

                program = Program(title=title, slug=slug, allow_rating=True, allow_comments=True)
                program_home.add_child(instance=program)

                # add category
                program.categories.add(category)

                # add program courses
                program.courses.get_or_create(program=program, course=course)

                program.save()
                program.save_revision().publish()

            # flex
            if not Flex.objects.filter(title=title, categories__name=category.name).exists():
                # fix duplicate flex
                slug = handle_duplicate_slug(Flex, slug)

                flex = Flex(title=title, slug=slug, allow_rating=True, allow_comments=True)
                flex_home.add_child(instance=flex)

                # add category
                flex.categories.add(category)

                # add flex courses
                flex.courses.get_or_create(flex=flex, course=course)

                flex.save()
                flex.save_revision().publish()

        self.stdout.write(
            self.style.SUCCESS(f"Data loaded successfully from {excel_file} to Course model")
        )
